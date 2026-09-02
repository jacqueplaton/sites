"""Import de CSV e export de CSV/XLSX."""

import io

from openpyxl import load_workbook

from app.crm.crud import criar_lead

CSV = """nome_empresa;categoria;cidade;estado;telefone;website;avaliacao;qtd_avaliacoes
Clínica Odonto Aurora;dentista;Campinas;SP;(19) 5555-0101;;4,8;143
Barbearia Dom Vito;barbearia;Campinas;SP;(19) 5555-0303;;4,7;210
Clinica Odonto Aurora ME;dentista;Campinas;SP;(19) 5555-0101;;4,8;143
;dentista;Campinas;SP;;;;
"""


def test_import_csv_com_duplicata_e_erro(cliente):
    arquivo = {"arquivo": ("leads.csv", CSV.encode("utf-8"), "text/csv")}
    relatorio = cliente.post("/api/import/csv", files=arquivo).json()
    assert relatorio["total_linhas"] == 4
    assert relatorio["importados"] == 2
    assert relatorio["duplicados"] == 1
    assert relatorio["erros"] == 1
    assert "telefone" in relatorio["detalhes_duplicados"][0]["motivo"]


def test_import_aceita_cabecalho_alternativo(cliente):
    csv = "Nome;Nicho;Cidade;Fone;Site\nPadaria Estrela;restaurante;Vinhedo;19 5555-2222;\n"
    relatorio = cliente.post(
        "/api/import/csv", files={"arquivo": ("x.csv", csv.encode("utf-8"), "text/csv")}
    ).json()
    assert relatorio["importados"] == 1
    lead = cliente.get("/api/leads").json()["itens"][0]
    assert lead["categoria"] == "restaurante" and lead["telefone"] == "19 5555-2222"


def test_import_sem_coluna_de_nome_falha(cliente):
    csv = "cidade;telefone\nCampinas;19 5555-0000\n"
    relatorio = cliente.post(
        "/api/import/csv", files={"arquivo": ("x.csv", csv.encode("utf-8"), "text/csv")}
    ).json()
    assert relatorio["importados"] == 0 and relatorio["erros"] == 1


def test_import_de_arquivo_vazio_da_400(cliente):
    assert cliente.post(
        "/api/import/csv", files={"arquivo": ("x.csv", b"", "text/csv")}
    ).status_code == 400


def test_import_nao_verifica_site_na_rede(cliente):
    """Import em massa não pode sair batendo em site nenhum."""
    csv = "nome_empresa;website\nEmpresa Teste;https://naodeveserchamado.example.com\n"
    cliente.post("/api/import/csv", files={"arquivo": ("x.csv", csv.encode("utf-8"), "text/csv")})
    lead = cliente.get("/api/leads").json()["itens"][0]
    assert lead["website_status"] == "NAO_VERIFICADO"
    assert lead["website_verificado_em"] is None


def test_export_csv_respeita_filtros(cliente, db):
    criar_lead(db, {"nome_empresa": "Odonto Aurora", "cidade": "Campinas"})
    criar_lead(db, {"nome_empresa": "Barbearia Dom Vito", "cidade": "Sorocaba"})
    resposta = cliente.get("/api/export/csv", params={"cidade": "Campinas"})
    assert resposta.status_code == 200
    texto = resposta.content.decode("utf-8-sig")
    assert "Odonto Aurora" in texto and "Dom Vito" not in texto
    assert texto.splitlines()[0].startswith("ID;Empresa")


def test_export_xlsx_abre_como_planilha(cliente, db):
    criar_lead(db, {"nome_empresa": "Odonto Aurora", "cidade": "Campinas", "score": 0})
    resposta = cliente.get("/api/export/xlsx")
    planilha = load_workbook(io.BytesIO(resposta.content))
    aba = planilha["Leads"]
    assert aba.cell(row=1, column=2).value == "Empresa"
    assert aba.cell(row=2, column=2).value == "Odonto Aurora"


def test_export_sem_leads_gera_arquivo_com_cabecalho(cliente):
    texto = cliente.get("/api/export/csv").content.decode("utf-8-sig")
    assert texto.strip().count("\n") == 0


def test_csv_exportado_pode_ser_reimportado(cliente, db):
    criar_lead(db, {
        "nome_empresa": "Odonto Aurora", "categoria": "dentista", "cidade": "Campinas",
        "telefone": "(19) 5555-0101", "avaliacao": 4.8, "qtd_avaliacoes": 143,
    })
    exportado = cliente.get("/api/export/csv").content
    # reimportar na mesma base tem de bloquear tudo por duplicidade
    relatorio = cliente.post(
        "/api/import/csv", files={"arquivo": ("saida.csv", exportado, "text/csv")}
    ).json()
    assert relatorio["total_linhas"] == 1
    assert relatorio["duplicados"] == 1
    assert relatorio["importados"] == 0
