"""Export de leads em CSV e XLSX, respeitando os filtros da tela."""

from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLUNAS: list[tuple[str, str]] = [
    ("id", "ID"),
    ("nome_empresa", "Empresa"),
    ("categoria", "Categoria"),
    ("subcategoria", "Subcategoria"),
    ("cidade", "Cidade"),
    ("estado", "Estado"),
    ("pais", "País"),
    ("endereco", "Endereço"),
    ("telefone", "Telefone"),
    ("website", "Website"),
    ("website_url_verificada", "Website verificado"),
    ("site_situacao", "Situação do site"),
    ("website_status", "Status do website"),
    ("website_confianca", "Confiança"),
    ("instagram", "Instagram"),
    ("facebook", "Facebook"),
    ("google_maps_url", "Google Maps"),
    ("avaliacao", "Nota"),
    ("qtd_avaliacoes", "Nº de avaliações"),
    ("horario", "Horário"),
    ("score", "Score"),
    ("faixa", "Faixa"),
    ("status", "Status"),
    ("tipo_abordagem", "Abordagem"),
    ("valor_proposta", "Valor da proposta"),
    ("motivo_perda", "Motivo da perda"),
    ("fonte", "Fonte"),
    ("data_coleta", "Data da coleta"),
    ("observacoes", "Observações"),
]


def _valor(lead: Any, campo: str) -> Any:
    valor = getattr(lead, campo, None)
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y %H:%M")
    return valor


def exportar_csv(leads: list[Any]) -> bytes:
    buffer = io.StringIO()
    escritor = csv.writer(buffer, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    escritor.writerow([rotulo for _, rotulo in COLUNAS])
    for lead in leads:
        escritor.writerow(["" if (v := _valor(lead, c)) is None else v for c, _ in COLUNAS])
    # BOM: o Excel em português abre o arquivo com acento correto.
    return buffer.getvalue().encode("utf-8-sig")


def exportar_xlsx(leads: list[Any]) -> bytes:
    planilha = Workbook()
    aba = planilha.active
    aba.title = "Leads"

    cabecalho_fill = PatternFill("solid", fgColor="1F2937")
    cabecalho_font = Font(color="FFFFFF", bold=True)
    for coluna, (_, rotulo) in enumerate(COLUNAS, start=1):
        celula = aba.cell(row=1, column=coluna, value=rotulo)
        celula.fill = cabecalho_fill
        celula.font = cabecalho_font
        celula.alignment = Alignment(vertical="center")

    for linha, lead in enumerate(leads, start=2):
        for coluna, (campo, _) in enumerate(COLUNAS, start=1):
            aba.cell(row=linha, column=coluna, value=_valor(lead, campo))

    for coluna, (campo, rotulo) in enumerate(COLUNAS, start=1):
        largura = max(
            len(rotulo),
            *(len(str(_valor(lead, campo) or "")) for lead in leads[:200]),
        ) if leads else len(rotulo)
        aba.column_dimensions[get_column_letter(coluna)].width = min(max(largura + 2, 10), 45)

    aba.freeze_panes = "A2"
    aba.auto_filter.ref = f"A1:{get_column_letter(len(COLUNAS))}{max(len(leads) + 1, 1)}"

    saida = io.BytesIO()
    planilha.save(saida)
    return saida.getvalue()


def nome_arquivo(extensao: str) -> str:
    return f"leads_{datetime.now().strftime('%Y%m%d_%H%M')}.{extensao}"
